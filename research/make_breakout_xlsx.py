"""정상 상승장(신고가 눌림) 표를 엑셀로. 낙폭 표와 같은 양식."""
import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUT = r"C:\Users\jangs_tjkt17a\Documents\stock_event_jarvis\output\나스닥_신고가눌림_수익률.xlsx"
NDX = """AAPL MSFT NVDA AMZN AVGO META TSLA GOOGL GOOG COST NFLX TMUS PLTR CSCO AMD LIN PEP
INTU ISRG TXN QCOM BKNG ADBE AMGN HON AMAT GILD PANW ADP VRTX MU LRCX ADI SBUX MELI KLAC INTC
CRWD CEG MDLZ CTAS PYPL CDNS SNPS MAR ORLY ABNB REGN FTNT ASML CSX WDAY TTD PDD ROP MNST AEP
NXPI DASH CHTR PCAR ADSK ROST FANG PAYX AZN KDP MRVL ODFL FAST EA CPRT VRSK IDXX EXC BKR CTSH
XEL CCEP KHC TEAM LULU ZS DXCM TTWO MCHP ON CDW GEHC WBD BIIB ILMN MDB ARM APP AXON""".split()

d = yf.download(NDX + ["QQQ"], period="10y", interval="1d", auto_adjust=True,
                group_by="ticker", threads=8, progress=False)
data = {}
for t in NDX + ["QQQ"]:
    try:
        df = d[t].dropna(how="all")[["Open", "High", "Low", "Close"]].dropna()
    except Exception:
        continue
    if len(df) >= 400:
        data[t] = df
Q = data.pop("QQQ")
QDD = Q["Close"] / Q["High"].rolling(252, min_periods=252).max() - 1.0
QMA = Q["Close"].rolling(200, min_periods=200).mean()
UP = set(Q.index[((Q["Close"] > QMA) & (QDD > -0.10)).fillna(False).values])
YEARS = (Q.index[-1] - Q.index[0]).days / 365.25
SPLIT = pd.Timestamp("2021-08-01")
HOLDS = [(60, "3개월"), (120, "6개월"), (250, "1년")]
WAITS = [(1, 3, "1~3일"), (3, 5, "3~5일"), (5, 10, "5~10일")]
DROPS = [(-0.04, -0.02, "2~4%"), (-0.06, -0.04, "4~6%"),
         (-0.10, -0.06, "6~10%"), (-0.15, -0.10, "10~15%")]

PRE = {}
for t, df in data.items():
    hi = df["High"].rolling(252, min_periods=252).max()
    PRE[t] = {"idx": df.index, "high": df["High"].values, "close": df["Close"].values,
              "hi": hi.values, "nh": (df["High"] >= hi).values,
              "ret": {lab: (df["Close"].shift(-(1 + h)) / df["Open"].shift(-1) - 1.0).values * 100
                      for h, lab in HOLDS}}


def signals(p, w0, w1, d0, d1):
    out, n = [], len(p["idx"])
    nh, cl, hg, h2 = p["nh"], p["close"], p["high"], p["hi"]
    i = 251
    while i < n:
        if not (np.isfinite(h2[i]) and nh[i]):
            i += 1
            continue
        peak, fired, j = hg[i], False, i
        for k in range(1, w1 + 1):
            j = i + k
            if j >= n or nh[j]:
                break
            peak = max(peak, hg[j])
            if k >= w0 and d0 <= cl[j] / peak - 1.0 <= d1:
                out.append(j)
                fired = True
                break
        i = (j if fired else i + 1)
    return out


def collect(w, db, lab, half=None):
    out = []
    for t, p in PRE.items():
        r = p["ret"][lab]
        for j in signals(p, w[0], w[1], db[0], db[1]):
            dt = p["idx"][j]
            if dt not in UP:
                continue
            if half == "a" and dt >= SPLIT:
                continue
            if half == "b" and dt < SPLIT:
                continue
            if np.isfinite(r[j]):
                out.append(r[j])
    return np.array(out)


def base(lab, half=None):
    out = []
    for t, p in PRE.items():
        r = p["ret"][lab]
        for i in range(251, len(p["idx"])):
            dt = p["idx"][i]
            if dt not in UP or not np.isfinite(p["hi"][i]) or not np.isfinite(r[i]):
                continue
            if half == "a" and dt >= SPLIT:
                continue
            if half == "b" and dt < SPLIT:
                continue
            out.append(r[i])
    return np.array(out)


BASE = {lab: base(lab) for _h, lab in HOLDS}
BA, BB = base("6개월", "a"), base("6개월", "b")

ROWS = []
for w in WAITS:
    for db in DROPS:
        cells, n = [], 0
        for _h, lab in HOLDS:
            v = collect(w, db, lab)
            n = max(n, len(v))
            cells.append((np.median(v), (v > 0).mean() * 100) if len(v) >= 100 else None)
        a, b = collect(w, db, "6개월", "a"), collect(w, db, "6개월", "b")
        da = (a > 0).mean() * 100 - (BA > 0).mean() * 100 if len(a) >= 50 else None
        dbv = (b > 0).mean() * 100 - (BB > 0).mean() * 100 if len(b) >= 50 else None
        if da is None or dbv is None:
            judge = "표본 부족"
        elif da > 0 and dbv > 0:
            judge = "양쪽 다 기준선 위"
        elif da <= 0 and dbv <= 0:
            judge = "양쪽 다 아래"
        else:
            judge = "한쪽만"
        ROWS.append((w[2], db[2], round(n / YEARS), cells, judge, da, dbv))

# ── 엑셀 ────────────────────────────────────────────────────────
F, RED, BLUE = "Arial", "FF0000", "0070C0"
thin = Side(style="thin", color="000000")
hair = Side(style="hair", color="808080")
med = Side(style="medium", color="000000")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
good = PatternFill("solid", fgColor="E8F3E2")
bad = PatternFill("solid", fgColor="FCE8E6")

wb = Workbook()
ws = wb.active
ws.title = "신고가눌림"

info = [("종  목", "나스닥100종목중 96종목", None),
        ("기  간", "2016년 8월 ~ 2026년 8월 ", "(10년)"),
        ("자  료", "Yahoo Finance 일봉, 배당 반영", None),
        ("지  수", "QQQ (나스닥100 지수)", None),
        ("장  세", "나스닥이 200일선 위 + 고점에서 10% 안쪽일 때만 ", "(전체 날의 70%)")]
for i, (k, v, red) in enumerate(info):
    r = 1 + i
    c = ws.cell(row=r, column=1, value=f"{k} :")
    c.font = Font(name=F, size=12)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
    b = ws.cell(row=r, column=2)
    if red:
        b.value = CellRichText(TextBlock(InlineFont(rFont=F, sz=12), v),
                               TextBlock(InlineFont(rFont=F, sz=12, color=RED), red))
    else:
        b.value = v
        b.font = Font(name=F, size=12)
    b.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22

HDR, SUB, START = 7, 8, 9
ws.merge_cells(start_row=HDR, start_column=1, end_row=SUB, end_column=1)
ws.cell(row=HDR, column=1, value="신고가 뒤\n며칠 기다려")
ws.merge_cells(start_row=HDR, start_column=2, end_row=SUB, end_column=2)
ws.cell(row=HDR, column=2, value="고점에서\n얼마나 눌렸을 때")
ws.merge_cells(start_row=HDR, start_column=3, end_row=SUB, end_column=3)
ws.cell(row=HDR, column=3, value="1년에\n몇 번 오나")
ws.merge_cells(start_row=HDR, start_column=4, end_row=HDR, end_column=6)
ws.cell(row=HDR, column=4).value = CellRichText(
    TextBlock(InlineFont(rFont=F, sz=12), "보유기간별  "),
    TextBlock(InlineFont(rFont=F, sz=12, color=RED), "수익율"),
    TextBlock(InlineFont(rFont=F, sz=12), " "),
    TextBlock(InlineFont(rFont=F, sz=12, color=BLUE), "(승률)"))
for i, lab in enumerate(["3개월", "6개월", "1년"]):
    ws.cell(row=SUB, column=4 + i, value=lab).font = Font(name=F, size=12)
ws.merge_cells(start_row=HDR, start_column=7, end_row=SUB, end_column=7)
ws.cell(row=HDR, column=7, value="교차 검증\n(앞 5년 / 뒤 5년)")
for r in (HDR, SUB):
    for c in range(1, 8):
        cell = ws.cell(row=r, column=c)
        if not cell.font or cell.font.sz != 12:
            cell.font = Font(name=F, size=12)
        cell.alignment = center

for i, (wl, dl, per_year, cells, judge, da, dbv) in enumerate(ROWS):
    r = START + i
    for c, val in ((1, wl), (2, dl), (3, f"{per_year}번")):
        cell = ws.cell(row=r, column=c, value=val)
        cell.font = Font(name=F, size=12)
        cell.alignment = center
    for k, cc in enumerate(cells):
        cell = ws.cell(row=r, column=4 + k)
        if cc is None:
            cell.value = "표본 부족"
            cell.font = Font(name=F, size=11, color="808080")
        else:
            bold = judge == "양쪽 다 기준선 위"
            cell.value = CellRichText(
                TextBlock(InlineFont(rFont=F, sz=12, b=bold, color=RED), f"{cc[0]:.1f}%"),
                TextBlock(InlineFont(rFont=F, sz=12, b=bold, color=BLUE), f" ({cc[1]:.1f}%)"))
        cell.alignment = center
    jc = ws.cell(row=r, column=7)
    jc.value = judge if da is None else f"{judge}\n{da:+.1f} / {dbv:+.1f}%p"
    jc.font = Font(name=F, size=11, bold=(judge == "양쪽 다 기준선 위"))
    jc.alignment = center
    if judge == "양쪽 다 기준선 위":
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = good
    elif judge == "양쪽 다 아래":
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = bad

END = START + len(ROWS) - 1
for r in range(HDR, END + 1):
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = Border(
            left=med if c in (1, 4, 7) else hair,
            right=med if c == 7 else hair,
            top=med if r == HDR else (thin if r == START else hair),
            bottom=med if r == END else hair)

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 12
for col in "DEF":
    ws.column_dimensions[col].width = 19
ws.column_dimensions["G"].width = 22
ws.row_dimensions[HDR].height = 34
ws.row_dimensions[SUB].height = 26
for r in range(START, END + 1):
    ws.row_dimensions[r].height = 32

n = END + 2
lines = [
    "읽는 법",
    "· 빨강 = 수익률(성적의 한가운데 값) · 파랑 = 승률(100번 사면 몇 번 이익)",
    f"· 비교 기준 — 정상 상승장 아무 날 아무 종목: 3개월 {np.median(BASE['3개월']):.1f}%({(BASE['3개월']>0).mean()*100:.1f}%) · "
    f"6개월 {np.median(BASE['6개월']):.1f}%({(BASE['6개월']>0).mean()*100:.1f}%) · "
    f"1년 {np.median(BASE['1년']):.1f}%({(BASE['1년']>0).mean()*100:.1f}%)",
    "· 초록 줄 = 앞 5년·뒤 5년 양쪽에서 기준선을 이긴 자리 (믿을 만함)",
    "· 붉은 줄 = 앞 5년·뒤 5년 양쪽에서 기준선보다 못한 자리 (쓰면 안 됨)",
    "· 교차 검증 칸의 숫자 = 6개월 승률이 그 시기 기준선보다 몇 %p 나은가",
    "",
    "지금 설명서는 '3~5일 기다려 4~6% 눌리면 매수'입니다 — 붉은 줄입니다.",
    "앞 5년 -0.2%p · 뒤 5년 -3.8%p로 양쪽 다 아무 종목이나 산 것보다 못했습니다.",
    "",
    f"매수는 신호 다음 거래일 시가, 매도는 정해진 거래일 뒤 종가. 3개월=60거래일 · 6개월=120거래일 · 1년=250거래일.",
    "신고가 하나당 한 번만 삽니다. 기다리는 중 새 신고가가 나오면 그 자리는 무효로 하고 새 신고가에서 다시 셉니다.",
]
for j, t in enumerate(lines):
    c = ws.cell(row=n + j, column=1, value=t)
    c.font = Font(name=F, size=11, bold=(j == 0 or t.startswith("지금 설명서")))

wb.save(OUT)
print("만들었습니다:", OUT)
print(f"\n기준선 6개월: 앞 5년 {(BA>0).mean()*100:.1f}% · 뒤 5년 {(BB>0).mean()*100:.1f}%")
for row in ROWS:
    print(f"  {row[0]:<7}{row[1]:<8}{row[2]:>4}번/년  {row[4]}")
