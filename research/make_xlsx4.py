"""캡처 화면 그대로 저장 — 주황 머리글 · 두 줄 칸 · 초록 강조."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

OUT = r"C:\Users\jangs_tjkt17a\Documents\stock_event_jarvis\output\나스닥_하락률별_수익률_최종.xlsx"
F, RED, BLUE = "Arial", "FF0000", "0070C0"

thin = Side(style="thin", color="000000")
hair = Side(style="hair", color="A6A6A6")
med = Side(style="medium", color="000000")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
HDR_FILL = PatternFill("solid", fgColor="FCE4D6")     # 주황
GREEN = PatternFill("solid", fgColor="E2EFDA")        # 초록

# 등급, 빈도, 사건수, [20~30%: 3M 6M 1Y], [30~50%: 3M 6M 1Y]
ROWS = [
    ("-6~-12%",  "7개월에 한 번", 25,
     (9.8, 71.5), (16.9, 74.6), (26.0, 73.6), (9.1, 67.7), (16.4, 69.5), (30.1, 71.6)),
    ("-12~-18%", "2.2년에 한 번", 7,
     (8.2, 63.7), (12.8, 67.0), (26.2, 74.1), (6.8, 57.2), (9.3, 59.6), (21.8, 62.5)),
    ("-18~-24%", "2.2년에 한 번", 4,
     (11.1, 69.5), (17.8, 71.7), (33.9, 85.4), (4.0, 56.9), (12.6, 63.1), (29.3, 76.2)),
    ("-24~-30%", "4.5년에 한 번", 2,
     (4.4, 62.7), (8.9, 68.0), (25.4, 89.5), (4.3, 61.0), (10.2, 63.8), (32.1, 87.0)),
    ("-30% 아래", "9년에 한 번", 1,
     (10.6, 81.7), (16.8, 92.9), (28.9, 96.8), (9.9, 71.7), (18.6, 82.4), (34.9, 91.7)),
]
# 초록 칠할 곳 — (줄번호, 칸번호들)
GREEN_CELLS = set()
for c in range(1, 9):
    GREEN_CELLS.add((0, c))                    # -6~-12% 줄 전체
for c in (3, 4, 5):
    GREEN_CELLS.add((2, c))                    # -18~-24% 의 20~30% 갈래
for c in (6, 7, 8):
    GREEN_CELLS.add((4, c))                    # -30% 아래 의 30~50% 갈래

wb = Workbook()
ws = wb.active
ws.title = "수익률표"

for i, (k, v, red) in enumerate([
        ("종  목", "나스닥100종목중 96종목", None),
        ("기  간", "2016년 8월 ~ 2026년 8월 ", "(10년)"),
        ("자  료", "Yahoo Finance 일봉, 배당 반영", None),
        ("지  수", "QQQ (나스닥100 지수)", None)]):
    r = 1 + i
    c = ws.cell(row=r, column=1, value=f"{k} :")
    c.font = Font(name=F, size=12)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    b = ws.cell(row=r, column=2)
    if red:
        b.value = CellRichText(TextBlock(InlineFont(rFont=F, sz=12), v),
                               TextBlock(InlineFont(rFont=F, sz=12, color=RED), red))
    else:
        b.value = v
        b.font = Font(name=F, size=12)
    b.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 24

HDR, SUB, START = 6, 7, 8

ws.merge_cells(start_row=HDR, start_column=1, end_row=SUB, end_column=1)
ws.cell(row=HDR, column=1, value="나스닥 고점\n대비 하락율").font = Font(name=F, size=12)
ws.merge_cells(start_row=HDR, start_column=2, end_row=SUB, end_column=2)
ws.cell(row=HDR, column=2, value="얼마나\n자주 오나\n(사건수)").font = Font(name=F, size=12)


def group_header(col, label):
    ws.merge_cells(start_row=HDR, start_column=col, end_row=HDR, end_column=col + 2)
    ws.cell(row=HDR, column=col).value = CellRichText(
        TextBlock(InlineFont(rFont=F, sz=12), f"{label} 빠진 종목  "),
        TextBlock(InlineFont(rFont=F, sz=12, color=RED), "수익율"),
        TextBlock(InlineFont(rFont=F, sz=12), " "),
        TextBlock(InlineFont(rFont=F, sz=12, color=BLUE), "(승률)"))


group_header(3, "20~30%")
group_header(6, "30~50%")
for i, lab in enumerate(["3개월", "6개월", "1년"] * 2):
    ws.cell(row=SUB, column=3 + i, value=lab).font = Font(name=F, size=12)

for r in (HDR, SUB):
    for c in range(1, 9):
        cell = ws.cell(row=r, column=c)
        cell.alignment = center
        cell.fill = HDR_FILL

for i, row in enumerate(ROWS):
    r = START + i
    a = ws.cell(row=r, column=1, value=row[0])
    a.font = Font(name=F, size=12)
    a.alignment = center
    b = ws.cell(row=r, column=2, value=f"{row[1]}\n({row[2]}번)")
    b.font = Font(name=F, size=12)
    b.alignment = center
    for k in range(6):
        col = 3 + k
        ret, win = row[3 + k]
        bold = (i, col) in GREEN_CELLS
        cell = ws.cell(row=r, column=col)
        cell.value = CellRichText(
            TextBlock(InlineFont(rFont=F, sz=12, b=bold, color=RED), f"{ret:.1f}%\n"),
            TextBlock(InlineFont(rFont=F, sz=12, b=bold, color=BLUE), f"({win:.1f}%)"))
        cell.alignment = center
    for c in range(1, 9):
        if (i, c) in GREEN_CELLS:
            ws.cell(row=r, column=c).fill = GREEN

END = START + len(ROWS) - 1
for r in range(HDR, END + 1):
    for c in range(1, 9):
        ws.cell(row=r, column=c).border = Border(
            left=med if c in (1, 3, 6) else hair,
            right=med if c == 8 else hair,
            top=med if r == HDR else (thin if r == START else hair),
            bottom=med if r == END else hair)

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 16
for col in "CDEFGH":
    ws.column_dimensions[col].width = 17
ws.row_dimensions[HDR].height = 38
ws.row_dimensions[SUB].height = 26
for r in range(START, END + 1):
    ws.row_dimensions[r].height = 42

n = END + 2
for j, t in enumerate([
    "읽는 법",
    "· 빨강 = 수익률(성적을 줄 세운 한가운데 값. 평균 아님) · 파랑 = 승률(100번 사면 몇 번 이익)",
    "· 초록 칸 = 그 갈래에서 가장 좋았던 자리",
    "· 비교 기준 — 아무 날에나 20~50% 빠진 종목을 사면 승률 3개월 61.4% · 6개월 65.4% · 1년 70.7%",
    "· 사건수가 진짜 한계입니다. -30% 아래는 코로나 1번, -24~-30%는 2번뿐입니다.",
    "· 실제로 쓸 수 있는 자리는 -6~-12% 하나입니다 (10년에 25번, 7개월에 한 번).",
    "",
    "매수는 신호 다음 거래일 시가, 매도는 정해진 거래일 뒤 종가. 3개월=60거래일 · 6개월=120거래일 · 1년=250거래일.",
]):
    c = ws.cell(row=n + j, column=1, value=t)
    c.font = Font(name=F, size=11, bold=(j == 0))

# ── 검증 시트 ───────────────────────────────────────────────────
v = wb.create_sheet("검증")
v["A1"] = "이 숫자를 얼마나 믿을 수 있나"
v["A1"].font = Font(name=F, size=14, bold=True)
v["A2"] = "'잰 횟수'가 커도 '사건수'가 적으면 그 몇 번을 잰 것입니다. 사건수가 진짜 한계입니다."
v["A2"].font = Font(name=F, size=10, italic=True, color="595959")

box = Border(left=thin, right=thin, top=thin, bottom=thin)
for i, h in enumerate(["나스닥 고점\n대비 하락율", "사건수\n(10년)", "잰 횟수\n20~30%",
                       "잰 횟수\n30~50%", "앞 5년\n수익률", "뒤 5년\n수익률", "판정"], start=1):
    c = v.cell(row=4, column=i, value=h)
    c.font = Font(name=F, size=11, bold=True)
    c.alignment = center
    c.border = box
    c.fill = HDR_FILL

for i, row in enumerate([
        ("-6~-12%", 25, 4479, 2897, 24.4, 8.3, "양쪽 다 이익 — 쓸 만함"),
        ("-12~-18%", 7, 2619, 2111, 29.2, -1.2, "뒤 5년에 무너짐"),
        ("-18~-24%", 4, 1782, 2090, 43.1, 3.5, "사건 4번뿐"),
        ("-24~-30%", 2, 1489, 2354, 56.3, 6.0, "사건 2번뿐"),
        ("-30% 아래", 1, 1270, 1709, None, None, "코로나 1번뿐 — 못 믿음")]):
    r = 5 + i
    for c, val in enumerate(row, start=1):
        cell = v.cell(row=r, column=c, value=val)
        cell.font = Font(name=F, size=11, bold=(i == 0))
        cell.alignment = center
        cell.border = box
        if c in (2, 3, 4):
            cell.number_format = "#,##0"
        if c in (5, 6):
            cell.number_format = '+0.0"%";-0.0"%";-'
        if i == 0:
            cell.fill = GREEN

for j, t in enumerate([
    "앞 5년 / 뒤 5년 = 10년을 2021년 8월에서 반으로 갈라 따로 잰 값(30~50% 빠진 종목 · 6개월 보유)",
    "뒤 5년이 모두 낮은 것은 앞 5년에 코로나 반등이 끼어 있어서입니다. 앞으로는 뒤 5년 쪽에 가깝다고 보는 편이 안전합니다.",
    "",
    "확인 방법",
    "· 자료를 새로 받아 대조했습니다 — 최대 차이 0.0002달러",
    "· 계산을 다른 방식으로 다시 짜서 일곱 칸 전부 0.05%p 안에서 일치를 확인했습니다",
    "· 신호를 만들 때 쓰는 값은 모두 그날까지의 것입니다(앞을 훔쳐보지 않음)",
]):
    c = v.cell(row=12 + j, column=1, value=t)
    c.font = Font(name=F, size=11, bold=t.startswith("확인 방법"))

v.column_dimensions["A"].width = 16
for col in "BCDEF":
    v.column_dimensions[col].width = 13
v.column_dimensions["G"].width = 26
v.row_dimensions[4].height = 34

wb.save(OUT)
print("저장했습니다:", OUT)
